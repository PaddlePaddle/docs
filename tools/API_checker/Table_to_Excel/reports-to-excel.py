#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 19 12:18:50 2019

@author: v_wanghao11
"""



from openpyxl import Workbook
from openpyxl import load_workbook

path = ''
wb = load_workbook(path+'all_api_table.xlsx')
wb_asignee = load_workbook(path+'APIAsignee.xlsx')
#print(wb.sheetnames)
ws=wb.active
rows=[]
for row in ws.iter_rows():
            rows.append(row)
#print rows[1][0].value
            
ws_asignee=wb_asignee.active
rows_asignee=[]
for row_asignee in ws_asignee.iter_rows():
            rows_asignee.append(row_asignee)
#print rows_asignee[1][1].value
       


report_cpu=open(path+"reports/report-cpu.txt",'r')
report_cuda=open(path+"reports/report_v100_cuda9.txt",'r')
debugging=open(path+"reports/debugging.txt",'w')


report_cpu_str=report_cpu.read()
report_cpu_array=report_cpu_str.split("FileName:")
#print(report_cpu_array[1])
for i in range(1,len(report_cpu_array)):
    report_cpu_api=report_cpu_array[i].split(".py:\n---------------------------",2)
    report_cpu_api_name=report_cpu_api[0]
    #print(type(report_cpu_api_name))
    report_cpu_api_result=report_cpu_api[1]
    #print(report_cpu_api_name+"#####"+report_cpu_api_result)
    for j in range(1, len(rows)):
        api_name_cmp=report_cpu_api_name
        par=report_cpu_api_name.find('(')
        if(par!=-1):
            api_name_cmp=api_name_cmp[:par]
        for temp in range(1,5):
            index_r=report_cpu_api_name.find('_'+str(temp))
            if(index_r!=-1):
                api_name_cmp=api_name_cmp[:index_r]
                break;
        if(api_name_cmp==rows[j][0].value):
            
            
            #the api in the api table is at row j
            
            #debugging.write("\n-----\n"+report_cpu_api_name+"===>"+rows[j][0].value+"\n-----\n");
            ws.cell(row=j+1, column=11, value=str(rows[j][10].value)+report_cpu_api_name+"\n========\n"+report_cpu_api_result)





report_cuda_str=report_cuda.read()
report_cuda_array=report_cuda_str.split("FileName:")

for i in range(1,len(report_cuda_array)):
    report_cuda_api=report_cuda_array[i].split(".py:\n---------------------------",2)
    report_cuda_api_name=report_cuda_api[0]
    
    report_cuda_api_result=report_cuda_api[1]
    #print(report_cpu_api_name+"#####"+report_cpu_api_result)
    for j in range(1, len(rows)):
        
        api_name_cmp=report_cuda_api_name
        par=report_cuda_api_name.find('(')
        if(par!=-1):
            api_name_cmp=api_name_cmp[:par]
        for temp in range(1,5):
            index_r=report_cuda_api_name.find('_'+str(temp))
            if(index_r!=-1):
                api_name_cmp=api_name_cmp[:index_r]
                break;
        #debugging.write("\n-----\n"+report_cuda_api_name+"===>"+api_name_cmp+"===>"+str(rows[j][0].value)+"\n-----\n");
        if(api_name_cmp==rows[j][0].value):
            ws.cell(row=j+1, column=12, value=str(rows[j][11].value)+report_cuda_api_name+"\n========\n"+report_cuda_api_result)


for k in range(1, len(rows_asignee)):
       if(rows_asignee[k][2].value==None):
           api_name_array_asignee = (rows_asignee[k][0].value).split(".",2)
       else:
           api_name_array_asignee = (rows_asignee[k][0].value).split(".",3)
       api_name_asignee=api_name_array_asignee[-1]
       api_name_asignee=api_name_asignee.replace(" ","")
       
       
       
       
       debugging.write("\n((((((((((\n%"+api_name_asignee+"\n)))))))))))))\n");
       #print("\n-----\n"+report_cpu_api_name+"\n-----\n")
       Found=False
       
       for j in range(1, len(rows)):
           
           debugging.write("\n'"+api_name_asignee+"'===>'"+str(rows[j][0].value)+"'\n");
           cmp_name=str(rows[j][0].value).replace(" ","")
           if(api_name_asignee.find(".")==-1):
               if(api_name_asignee==cmp_name):
                       
                      #debugging.write("\n-----\napi_name_asignee: '"+api_name_asignee+"'\n");
                      #debugging.write("api_name_in_table: '"+str(rows[j][0].value)+"'\n-----\n");
                      Found=True
                      print(rows_asignee[k][1].value)
                      ws.cell(row=j+1, column=13, value=rows_asignee[k][1].value)
           else:
                debugging.write("\n-#-#-#-#-\napi_name_asignee: $'"+api_name_asignee[:api_name_asignee.find(".")]+"'\n");
                debugging.write("api_name_in_table: $'"+str(rows[j][0].value)[:str(rows[j][0].value).find(".")]+"'\n-#-#-#-#-\n");
                   
                if(api_name_asignee[:api_name_asignee.find(".")]==str(rows[j][0].value)[:str(rows[j][0].value).find(".")]):
                        debugging.write("\n-?--?-?-?--\napi_name_asignee: *'"+api_name_asignee[:api_name_asignee.find(".")]+"'\n");
                        debugging.write("api_name_in_table: '"+str(rows[j][0].value)+"'\n-?--?-?-?--\n");
                        Found=True
                        print(rows_asignee[k][1].value)
                        
                        ws.cell(row=j+1, column=13, value=rows_asignee[k][1].value)
                        
                        if(rows_asignee[k][1].value==None):
                            ws.cell(row=j+1, column=13, value="API Unassigned")
       
       if(Found==False):
                  debugging.write("\n*****\napi_name_asignee: "+api_name_asignee+"\n******\n");
                 



wb.save(path+'all_api_table_after.xlsx')

report_cpu.close();
report_cuda.close();
debugging.close();


